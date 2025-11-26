#!/usr/bin/env python3
"""
PMP MSP Dashboard Generator - Phase 191
Comprehensive Excel reports for MSP multi-tenant environment

Features:
- Executive summary across all 30 organizations
- Per-organization health scorecards
- Critical systems detail
- Policy overview
- OS distribution analysis
- Top 10 critical organizations deep dive

Requirements:
    pip3 install openpyxl

Author: Patch Manager Plus API Specialist Agent
Date: 2025-11-26
Version: 1.0 (Phase 191)
"""

import sqlite3
from pathlib import Path
from typing import Optional, List, Dict, Tuple
from datetime import datetime
import logging

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ Error: openpyxl not installed")
    print("Install with: pip3 install openpyxl")
    import sys
    sys.exit(1)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class PMPMSPDashboard:
    """
    Comprehensive MSP dashboard generator for PMP

    Creates multi-sheet Excel workbook with:
    - Executive summary (all organizations)
    - Organization details (per-org breakdown)
    - Critical systems (high-risk focus)
    - Policy overview (deployment policies)
    - OS distribution (license planning)
    """

    def __init__(self, db_path: Optional[Path] = None):
        """Initialize dashboard generator"""
        if db_path is None:
            db_path = Path.home() / ".maia/databases/intelligence/pmp_config.db"

        self.db_path = Path(db_path)

        if not self.db_path.exists():
            raise FileNotFoundError(f"Database not found: {self.db_path}")

        # Style definitions
        self.header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True, size=11)

        self.critical_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        self.critical_font = Font(color="FFFFFF", bold=True)

        self.warning_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        self.success_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")

        self.border_thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def generate_msp_dashboard(self, output_dir: Optional[Path] = None) -> Path:
        """
        Generate complete MSP dashboard Excel report

        Args:
            output_dir: Output directory (default: ~/work_projects/pmp_reports/)

        Returns:
            Path to generated Excel file
        """
        if output_dir is None:
            output_dir = Path.home() / "work_projects/pmp_reports"

        output_dir.mkdir(parents=True, exist_ok=True)

        # Generate filename
        timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        filename = f"PMP_MSP_Dashboard_{timestamp}.xlsx"
        output_path = output_dir / filename

        print(f"\n{'='*70}")
        print(f"📊 PMP MSP DASHBOARD GENERATOR")
        print(f"{'='*70}\n")

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create worksheets
        print("  → Executive Summary...")
        self._create_executive_summary(wb)

        print("  → Organization Details...")
        self._create_organization_details(wb)

        print("  → Critical Systems...")
        self._create_critical_systems(wb)

        print("  → OS Distribution...")
        self._create_os_distribution(wb)

        print("  → Policy Overview...")
        self._create_policy_overview(wb)

        print("  → Top 10 Critical Orgs...")
        self._create_top_critical_orgs(wb)

        # Save workbook
        wb.save(output_path)

        print(f"\n{'='*70}")
        print(f"✅ DASHBOARD GENERATED")
        print(f"{'='*70}")
        print(f"📄 File: {output_path}")
        print(f"📊 Size: {output_path.stat().st_size / 1024:.1f} KB")
        print(f"{'='*70}\n")

        logger.info(f"MSP dashboard generated: {output_path}")

        return output_path

    def _create_executive_summary(self, wb: Workbook):
        """Executive Summary - All Organizations Overview"""
        ws = wb.create_sheet("Executive Summary")

        # Get summary data
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        # Organization summary
        cursor.execute("""
            SELECT
                branch_office_name as organization,
                COUNT(*) as total_systems,
                SUM(CASE WHEN resource_health_status = 1 THEN 1 ELSE 0 END) as healthy,
                SUM(CASE WHEN resource_health_status = 2 THEN 1 ELSE 0 END) as moderate,
                SUM(CASE WHEN resource_health_status = 3 THEN 1 ELSE 0 END) as high_risk,
                ROUND(100.0 * SUM(CASE WHEN resource_health_status = 3 THEN 1 ELSE 0 END) / COUNT(*), 1) as risk_pct
            FROM systems
            GROUP BY branch_office_name
            ORDER BY risk_pct DESC, total_systems DESC
        """)

        orgs = cursor.fetchall()

        # Totals
        cursor.execute("""
            SELECT
                COUNT(DISTINCT branch_office_name) as total_orgs,
                COUNT(*) as total_systems,
                SUM(CASE WHEN resource_health_status = 1 THEN 1 ELSE 0 END) as healthy,
                SUM(CASE WHEN resource_health_status = 3 THEN 1 ELSE 0 END) as high_risk
            FROM systems
        """)

        totals = cursor.fetchone()
        conn.close()

        # Header
        ws['A1'] = "ManageEngine PMP - MSP Executive Dashboard"
        ws['A1'].font = Font(size=16, bold=True, color="2F5496")
        ws.merge_cells('A1:F1')

        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(size=10, italic=True)
        ws.merge_cells('A2:F2')

        # KPI Section
        row = 4
        ws[f'A{row}'] = "Key Metrics"
        ws[f'A{row}'].font = Font(size=14, bold=True)

        row += 1
        kpis = [
            ("Total Organizations", totals[0], None),
            ("Total Systems", totals[1], None),
            ("Healthy Systems", totals[2], self.success_fill),
            ("High Risk Systems", totals[3], self.critical_fill if totals[3] > 0 else None),
        ]

        for label, value, fill in kpis:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = Font(size=12, bold=True)

            if fill:
                ws[f'B{row}'].fill = fill
                if fill == self.critical_fill:
                    ws[f'B{row}'].font = Font(size=12, bold=True, color="FFFFFF")

            row += 1

        # Organization table
        row += 2
        ws[f'A{row}'] = "All Organizations"
        ws[f'A{row}'].font = Font(size=14, bold=True)

        row += 1
        headers = ['Organization', 'Total Systems', 'Healthy', 'Moderate', 'High Risk', 'Risk %']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border_thin

        row += 1
        for org_data in orgs:
            org, total, healthy, moderate, high_risk, risk_pct = org_data

            ws[f'A{row}'] = org
            ws[f'B{row}'] = total
            ws[f'C{row}'] = healthy
            ws[f'D{row}'] = moderate
            ws[f'E{row}'] = high_risk
            ws[f'F{row}'] = f"{risk_pct}%"

            # Risk % color coding
            if risk_pct >= 70:
                ws[f'F{row}'].fill = self.critical_fill
                ws[f'F{row}'].font = Font(bold=True, color="FFFFFF")
            elif risk_pct >= 40:
                ws[f'F{row}'].fill = self.warning_fill
                ws[f'F{row}'].font = Font(bold=True)

            # Apply borders
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = self.border_thin

            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 10

    def _create_organization_details(self, wb: Workbook):
        """Organization Details - Per-Org Breakdown"""
        ws = wb.create_sheet("Organization Details")

        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        # Get per-organization details including OS breakdown
        cursor.execute("""
            SELECT
                branch_office_name,
                COUNT(*) as total_systems,
                SUM(CASE WHEN resource_health_status = 1 THEN 1 ELSE 0 END) as healthy,
                SUM(CASE WHEN resource_health_status = 3 THEN 1 ELSE 0 END) as high_risk,
                SUM(CASE WHEN computer_live_status = 1 THEN 1 ELSE 0 END) as online,
                SUM(CASE WHEN computer_live_status = 0 THEN 1 ELSE 0 END) as offline,
                COUNT(DISTINCT os_name) as os_types,
                MAX(agent_last_contact_time) as last_contact
            FROM systems
            GROUP BY branch_office_name
            ORDER BY total_systems DESC
        """)

        orgs = cursor.fetchall()
        conn.close()

        # Header
        ws['A1'] = "Organization Details"
        ws['A1'].font = Font(size=14, bold=True)

        row = 3
        headers = ['Organization', 'Total', 'Healthy', 'High Risk', 'Online', 'Offline', 'OS Types', 'Last Contact']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border_thin

        row += 1
        for org_data in orgs:
            org, total, healthy, high_risk, online, offline, os_types, last_contact = org_data

            ws[f'A{row}'] = org
            ws[f'B{row}'] = total
            ws[f'C{row}'] = healthy
            ws[f'D{row}'] = high_risk
            ws[f'E{row}'] = online
            ws[f'F{row}'] = offline
            ws[f'G{row}'] = os_types
            ws[f'H{row}'] = datetime.fromtimestamp(last_contact/1000).strftime('%Y-%m-%d') if last_contact else 'N/A'

            # High risk highlighting
            if high_risk > 0:
                ws[f'D{row}'].fill = self.critical_fill
                ws[f'D{row}'].font = Font(bold=True, color="FFFFFF")

            # Offline highlighting
            if offline > 0:
                ws[f'F{row}'].fill = self.warning_fill
                ws[f'F{row}'].font = Font(bold=True)

            # Apply borders
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = self.border_thin

            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 45
        for col in ['B', 'C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 12
        ws.column_dimensions['H'].width = 14

    def _create_critical_systems(self, wb: Workbook):
        """Critical Systems - All High Risk Systems"""
        ws = wb.create_sheet("Critical Systems")

        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        # Get all high-risk systems
        cursor.execute("""
            SELECT
                branch_office_name,
                resource_name,
                ip_address,
                os_name,
                agent_version,
                agent_last_contact_time,
                last_scan_time
            FROM systems
            WHERE resource_health_status = 3
            ORDER BY branch_office_name, resource_name
        """)

        systems = cursor.fetchall()
        conn.close()

        # Header
        ws['A1'] = f"Critical Systems - High Risk ({len(systems)} systems)"
        ws['A1'].font = Font(size=14, bold=True, color="FF0000")

        if not systems:
            ws['A3'] = "✅ No high-risk systems found!"
            ws['A3'].font = Font(size=12, bold=True, color="00B050")
            return

        row = 3
        headers = ['Organization', 'System Name', 'IP Address', 'OS', 'Agent Version', 'Last Contact', 'Last Scan']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.critical_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border_thin

        row += 1
        for sys_data in systems:
            org, name, ip, os, agent_ver, last_contact, last_scan = sys_data

            ws[f'A{row}'] = org
            ws[f'B{row}'] = name
            ws[f'C{row}'] = ip if ip else 'N/A'
            ws[f'D{row}'] = os
            ws[f'E{row}'] = agent_ver if agent_ver else 'N/A'
            ws[f'F{row}'] = datetime.fromtimestamp(last_contact/1000).strftime('%Y-%m-%d %H:%M') if last_contact else 'N/A'
            ws[f'G{row}'] = datetime.fromtimestamp(last_scan/1000).strftime('%Y-%m-%d') if last_scan else 'N/A'

            # Apply borders
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = self.border_thin

            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 12

    def _create_os_distribution(self, wb: Workbook):
        """OS Distribution - License Planning"""
        ws = wb.create_sheet("OS Distribution")

        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        # Get OS distribution
        cursor.execute("""
            SELECT
                COALESCE(os_name, 'Unknown') as os,
                COUNT(*) as count,
                ROUND(100.0 * COUNT(*) / (SELECT COUNT(*) FROM systems), 1) as percentage
            FROM systems
            GROUP BY os_name
            ORDER BY count DESC
        """)

        os_data = cursor.fetchall()
        conn.close()

        # Header
        ws['A1'] = "Operating System Distribution"
        ws['A1'].font = Font(size=14, bold=True)

        row = 3
        headers = ['Operating System', 'System Count', 'Percentage']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border_thin

        row += 1
        for os, count, pct in os_data:
            ws[f'A{row}'] = os
            ws[f'B{row}'] = count
            ws[f'C{row}'] = f"{pct}%"

            # Apply borders
            for col in range(1, 4):
                ws.cell(row=row, column=col).border = self.border_thin

            row += 1

        # Create pie chart
        chart = PieChart()
        chart.title = "OS Distribution"

        labels = Reference(ws, min_col=1, min_row=4, max_row=row-1)
        data = Reference(ws, min_col=2, min_row=3, max_row=row-1)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)

        ws.add_chart(chart, "E3")

        # Column widths
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 12

    def _create_policy_overview(self, wb: Workbook):
        """Policy Overview - Deployment Policies"""
        ws = wb.create_sheet("Policy Overview")

        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        # Get deployment policies
        cursor.execute("""
            SELECT
                config_name,
                platform_name,
                config_status
            FROM deployment_policies
            WHERE snapshot_id = (SELECT MAX(snapshot_id) FROM snapshots WHERE status = 'success')
            ORDER BY config_name
        """)

        policies = cursor.fetchall()
        conn.close()

        # Header
        ws['A1'] = f"Deployment Policies ({len(policies)} policies)"
        ws['A1'].font = Font(size=14, bold=True)

        if not policies:
            ws['A3'] = "No deployment policies found"
            return

        row = 3
        headers = ['Policy Name', 'Platform', 'Status']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border_thin

        row += 1
        for name, platform, status in policies:
            ws[f'A{row}'] = name
            ws[f'B{row}'] = platform
            ws[f'C{row}'] = status

            # Apply borders
            for col in range(1, 4):
                ws.cell(row=row, column=col).border = self.border_thin

            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 70
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15

    def _create_top_critical_orgs(self, wb: Workbook):
        """Top 10 Critical Organizations - Deep Dive"""
        ws = wb.create_sheet("Top 10 Critical Orgs")

        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        # Get top 10 most critical organizations
        cursor.execute("""
            SELECT
                branch_office_name,
                COUNT(*) as total_systems,
                SUM(CASE WHEN resource_health_status = 3 THEN 1 ELSE 0 END) as high_risk,
                ROUND(100.0 * SUM(CASE WHEN resource_health_status = 3 THEN 1 ELSE 0 END) / COUNT(*), 1) as risk_pct,
                SUM(CASE WHEN computer_live_status = 0 THEN 1 ELSE 0 END) as offline
            FROM systems
            GROUP BY branch_office_name
            HAVING high_risk > 0
            ORDER BY risk_pct DESC, high_risk DESC
            LIMIT 10
        """)

        top_orgs = cursor.fetchall()
        conn.close()

        # Header
        ws['A1'] = "Top 10 Critical Organizations - Immediate Action Required"
        ws['A1'].font = Font(size=14, bold=True, color="FF0000")

        row = 3
        headers = ['Organization', 'Total Systems', 'High Risk', 'Risk %', 'Offline']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.critical_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border_thin

        row += 1
        for org, total, high_risk, risk_pct, offline in top_orgs:
            ws[f'A{row}'] = org
            ws[f'B{row}'] = total
            ws[f'C{row}'] = high_risk
            ws[f'D{row}'] = f"{risk_pct}%"
            ws[f'E{row}'] = offline

            # Critical highlighting
            ws[f'C{row}'].fill = self.critical_fill
            ws[f'C{row}'].font = Font(bold=True, color="FFFFFF")

            ws[f'D{row}'].fill = self.critical_fill
            ws[f'D{row}'].font = Font(bold=True, color="FFFFFF")

            # Apply borders
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = self.border_thin

            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10


def main():
    """CLI interface"""
    import argparse

    parser = argparse.ArgumentParser(description="PMP MSP Dashboard Generator")
    parser.add_argument('--output', type=str,
                       help='Output directory (default: ~/work_projects/pmp_reports/)')
    parser.add_argument('--db', type=str,
                       help='Path to SQLite database')

    args = parser.parse_args()

    db_path = Path(args.db) if args.db else None
    output_dir = Path(args.output) if args.output else None

    try:
        dashboard = PMPMSPDashboard(db_path=db_path)
        report_path = dashboard.generate_msp_dashboard(output_dir=output_dir)

        print(f"✅ Open with: open '{report_path}'")

    except Exception as e:
        print(f"❌ Dashboard generation failed: {e}")
        import traceback
        traceback.print_exc()
        import sys
        sys.exit(1)


if __name__ == '__main__':
    main()
