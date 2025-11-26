#!/usr/bin/env python3
"""
PMP Policy Complete Export - Phase 191
Export all policies with complete details to Excel

Features:
- Parse all deployment policies from JSON
- Extract all fields (template ID, name, dates, creator, settings)
- Export to comprehensive Excel spreadsheet
- Include all metadata and configuration details

Author: Patch Manager Plus API Specialist Agent
Date: 2025-11-26
Version: 1.0 (Phase 191)
"""

import sqlite3
import json
from pathlib import Path
from typing import Optional, List, Dict
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ Error: openpyxl not installed")
    print("Install with: pip3 install openpyxl")
    import sys
    sys.exit(1)


class PMPPolicyExport:
    """Export all PMP policies with complete details to Excel"""

    def __init__(self, db_path: Optional[Path] = None):
        """Initialize policy exporter"""
        if db_path is None:
            db_path = Path.home() / ".maia/databases/intelligence/pmp_config.db"

        self.db_path = Path(db_path)

        if not self.db_path.exists():
            raise FileNotFoundError(f"Database not found: {self.db_path}")

        # Style definitions
        self.header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True, size=11)
        self.border_thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def export_all_policies(self, output_dir: Optional[Path] = None) -> Path:
        """
        Export all policies with complete details to Excel

        Returns:
            Path to generated Excel file
        """
        if output_dir is None:
            output_dir = Path.home() / "work_projects/pmp_reports"

        output_dir.mkdir(parents=True, exist_ok=True)

        print(f"\n{'='*70}")
        print(f"📤 PMP POLICY COMPLETE EXPORT")
        print(f"{'='*70}\n")

        # Load all policies from all snapshots
        print("  → Loading policies from database...")
        policies = self._load_all_policies()

        print(f"     Found: {len(policies)} total policies")

        # Generate filename
        timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        filename = f"PMP_Policies_Complete_Export_{timestamp}.xlsx"
        output_path = output_dir / filename

        # Create Excel report
        print("  → Creating Excel export...")
        self._create_excel_export(output_path, policies)

        print(f"\n{'='*70}")
        print(f"✅ EXPORT COMPLETE")
        print(f"{'='*70}")
        print(f"📄 File: {output_path}")
        print(f"📊 Size: {output_path.stat().st_size / 1024:.1f} KB")
        print(f"📋 Policies: {len(policies)}")
        print(f"{'='*70}\n")

        return output_path

    def _load_all_policies(self) -> List[Dict]:
        """Load ALL deployment policies from database (all snapshots)"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        # Get all policies from all snapshots
        cursor.execute("""
            SELECT
                snapshot_id,
                settings_json
            FROM deployment_policies
            ORDER BY snapshot_id DESC, policy_id
        """)

        policies = []
        for row in cursor.fetchall():
            try:
                policy = json.loads(row[1])
                policy['_snapshot_id'] = row[0]  # Add snapshot tracking
                policies.append(policy)
            except json.JSONDecodeError:
                continue

        conn.close()
        return policies

    def _create_excel_export(self, output_path: Path, policies: List[Dict]):
        """Create comprehensive Excel export"""
        wb = Workbook()
        wb.remove(wb.active)

        # Sheet 1: Policy Overview (key fields)
        self._create_overview_sheet(wb, policies)

        # Sheet 2: Complete Details (all fields)
        self._create_complete_details_sheet(wb, policies)

        # Sheet 3: Raw JSON (for reference)
        self._create_json_sheet(wb, policies)

        wb.save(output_path)

    def _create_overview_sheet(self, wb: Workbook, policies: List[Dict]):
        """Policy Overview - Key Fields"""
        ws = wb.create_sheet("Policy Overview")

        # Header
        ws['A1'] = f"Deployment Policies - Overview ({len(policies)} policies)"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:M1')

        ws['A2'] = f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(size=10, italic=True)
        ws.merge_cells('A2:M2')

        # Headers
        row = 4
        headers = [
            'Snapshot', 'Template ID', 'Template Name', 'Created By', 'Created Date',
            'Modified Date', 'Customer ID', 'User ID', 'Set as Default',
            'Is Alive', 'Platform ID', 'Modified By User', 'Modified By Name'
        ]

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = self.border_thin

        # Data rows
        row += 1
        for policy in policies:
            # Extract key fields
            snapshot_id = policy.get('_snapshot_id', 'N/A')
            template_id = policy.get('template_id', 'N/A')
            template_name = policy.get('template_name', 'Unnamed')
            first_name = policy.get('first_name', 'Unknown')
            creation_time = policy.get('creation_time', 0)
            modified_time = policy.get('modified_time', 0)
            customer_id = policy.get('customer_id', 'N/A')
            user_id = policy.get('user_id', 'N/A')
            set_as_default = policy.get('set_as_default', False)
            is_alive = policy.get('is_template_alive', True)
            platform_id = policy.get('scriptaction.platform_id', 'N/A')
            modified_by_user = policy.get('usertodeploytemplaterel.modified_by', 'N/A')
            modified_by_name = policy.get('modifieduser.first_name', 'Unknown')

            # Write data
            ws.cell(row=row, column=1).value = snapshot_id
            ws.cell(row=row, column=2).value = str(template_id)
            ws.cell(row=row, column=3).value = template_name
            ws.cell(row=row, column=4).value = first_name
            ws.cell(row=row, column=5).value = datetime.fromtimestamp(creation_time/1000).strftime('%Y-%m-%d %H:%M') if creation_time > 0 else 'N/A'
            ws.cell(row=row, column=6).value = datetime.fromtimestamp(modified_time/1000).strftime('%Y-%m-%d %H:%M') if modified_time > 0 else 'N/A'
            ws.cell(row=row, column=7).value = str(customer_id)
            ws.cell(row=row, column=8).value = str(user_id)
            ws.cell(row=row, column=9).value = 'Yes' if set_as_default else 'No'
            ws.cell(row=row, column=10).value = 'Yes' if is_alive else 'No'
            ws.cell(row=row, column=11).value = str(platform_id)
            ws.cell(row=row, column=12).value = str(modified_by_user)
            ws.cell(row=row, column=13).value = modified_by_name

            # Apply borders
            for col in range(1, 14):
                ws.cell(row=row, column=col).border = self.border_thin
                ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical='top')

            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 60
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 10
        ws.column_dimensions['K'].width = 12
        ws.column_dimensions['L'].width = 15
        ws.column_dimensions['M'].width = 15

    def _create_complete_details_sheet(self, wb: Workbook, policies: List[Dict]):
        """Complete Details - All Fields Extracted"""
        ws = wb.create_sheet("Complete Details")

        # Header
        ws['A1'] = f"Deployment Policies - Complete Details ({len(policies)} policies)"
        ws['A1'].font = Font(size=14, bold=True)

        # Collect all unique keys across all policies
        all_keys = set()
        for policy in policies:
            all_keys.update(policy.keys())

        all_keys = sorted([k for k in all_keys if not k.startswith('_')])  # Exclude internal fields

        # Headers
        row = 3
        headers = ['Snapshot ID', 'Template Name'] + all_keys

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = self.border_thin

        # Data rows
        row += 1
        for policy in policies:
            snapshot_id = policy.get('_snapshot_id', 'N/A')
            template_name = policy.get('template_name', 'Unnamed')

            ws.cell(row=row, column=1).value = str(snapshot_id)
            ws.cell(row=row, column=2).value = template_name

            # Write all fields
            for col_idx, key in enumerate(all_keys, start=3):
                value = policy.get(key, '')

                # Format value
                if isinstance(value, bool):
                    cell_value = 'Yes' if value else 'No'
                elif isinstance(value, (int, float)) and key.endswith('_time') and value > 1000000000:
                    # Unix timestamp
                    try:
                        cell_value = datetime.fromtimestamp(value/1000).strftime('%Y-%m-%d %H:%M')
                    except:
                        cell_value = str(value)
                elif value == '--':
                    cell_value = 'N/A'
                else:
                    cell_value = str(value) if value is not None else ''

                ws.cell(row=row, column=col_idx).value = cell_value
                ws.cell(row=row, column=col_idx).border = self.border_thin
                ws.cell(row=row, column=col_idx).alignment = Alignment(wrap_text=True, vertical='top')

            row += 1

        # Auto-size columns (with max width)
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = min(50, max(12, len(headers[col_idx-1]) + 2))

    def _create_json_sheet(self, wb: Workbook, policies: List[Dict]):
        """Raw JSON - Full Policy Data"""
        ws = wb.create_sheet("Raw JSON")

        # Header
        ws['A1'] = f"Deployment Policies - Raw JSON ({len(policies)} policies)"
        ws['A1'].font = Font(size=14, bold=True)

        ws['A2'] = "Complete JSON data for each policy (for reference/debugging)"
        ws['A2'].font = Font(size=10, italic=True)

        # Headers
        row = 4
        ws['A4'] = 'Snapshot ID'
        ws['B4'] = 'Template Name'
        ws['C4'] = 'Complete JSON'

        for col in ['A', 'B', 'C']:
            ws[f'{col}4'].font = self.header_font
            ws[f'{col}4'].fill = self.header_fill
            ws[f'{col}4'].border = self.border_thin

        # Data
        row = 5
        for policy in policies:
            snapshot_id = policy.get('_snapshot_id', 'N/A')
            template_name = policy.get('template_name', 'Unnamed')

            # Remove internal fields for clean JSON
            clean_policy = {k: v for k, v in policy.items() if not k.startswith('_')}
            json_str = json.dumps(clean_policy, indent=2)

            ws[f'A{row}'] = str(snapshot_id)
            ws[f'B{row}'] = template_name
            ws[f'C{row}'] = json_str

            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].border = self.border_thin
                ws[f'{col}{row}'].alignment = Alignment(wrap_text=True, vertical='top')

            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 100


def main():
    """CLI interface"""
    import argparse

    parser = argparse.ArgumentParser(description="PMP Policy Complete Export")
    parser.add_argument('--output', type=str,
                       help='Output directory (default: ~/work_projects/pmp_reports/)')
    parser.add_argument('--db', type=str,
                       help='Path to SQLite database')

    args = parser.parse_args()

    db_path = Path(args.db) if args.db else None
    output_dir = Path(args.output) if args.output else None

    try:
        exporter = PMPPolicyExport(db_path=db_path)
        report_path = exporter.export_all_policies(output_dir=output_dir)

        print(f"✅ Open with: open '{report_path}'")

    except Exception as e:
        print(f"❌ Export failed: {e}")
        import traceback
        traceback.print_exc()
        import sys
        sys.exit(1)


if __name__ == '__main__':
    main()
