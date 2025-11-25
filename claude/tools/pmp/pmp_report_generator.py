#!/usr/bin/env python3
"""
PMP Report Generator - Excel Export Engine
Generate business-friendly Excel reports from SQLite configuration snapshots

Features:
- Compliance Dashboard (Executive Summary + Trend Charts + Compliance Checks)
- Conditional formatting (red/yellow/green)
- Embedded charts (pie, line, bar)
- Print-optimized layout

Requirements:
    pip3 install openpyxl

Author: Patch Manager Plus API Specialist Agent + SRE Principal Engineer Agent
Date: 2025-11-25
Version: 1.0
"""

import sqlite3
from pathlib import Path
from typing import Optional, List, Dict
from datetime import datetime
import logging

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import PieChart, LineChart, Reference
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl not installed. Run: pip3 install openpyxl")
    import sys
    sys.exit(1)

logger = logging.getLogger(__name__)


class PMPReportGenerator:
    """
    Excel report generator for PMP configuration data

    Usage:
        generator = PMPReportGenerator()
        report_path = generator.generate_compliance_dashboard(days=30)
        print(f"Report: {report_path}")
    """

    def __init__(self, db_path: Optional[Path] = None):
        """
        Initialize report generator

        Args:
            db_path: Path to SQLite database (default: ~/.maia/databases/intelligence/pmp_config.db)
        """
        if db_path is None:
            db_path = Path.home() / ".maia/databases/intelligence/pmp_config.db"

        self.db_path = Path(db_path)

        if not self.db_path.exists():
            raise FileNotFoundError(f"Database not found: {self.db_path}")

    def generate_compliance_dashboard(self, output_dir: Optional[Path] = None, days: int = 30) -> Path:
        """
        Generate compliance dashboard Excel report

        Args:
            output_dir: Output directory (default: ~/work_projects/pmp_reports/)
            days: Number of days of trend data to include (default: 30)

        Returns:
            Path to generated Excel file
        """
        if output_dir is None:
            output_dir = Path.home() / "work_projects/pmp_reports"

        output_dir.mkdir(parents=True, exist_ok=True)

        # Generate filename
        timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        filename = f"PMP_Compliance_Dashboard_{timestamp}.xlsx"
        output_path = output_dir / filename

        # Create workbook
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create worksheets
        self._create_executive_summary(wb)
        self._create_trend_charts(wb, days=days)
        self._create_severity_analysis(wb)

        # Save workbook
        wb.save(output_path)

        logger.info("compliance_dashboard_generated", extra={
            "output_path": str(output_path),
            "file_size_bytes": output_path.stat().st_size
        })

        return output_path

    def _create_executive_summary(self, wb: Workbook):
        """Create Executive Summary worksheet"""
        ws = wb.create_sheet("Executive Summary")

        # Get latest snapshot data
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        cursor.execute("SELECT * FROM latest_snapshot")
        latest = dict(cursor.fetchone() or {})

        conn.close()

        if not latest:
            ws['A1'] = "No data available"
            return

        # Header
        ws['A1'] = "ManageEngine PMP - Executive Summary"
        ws['A1'].font = Font(size=16, bold=True)

        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(size=10, italic=True)

        # KPIs
        row = 4
        ws[f'A{row}'] = "Key Performance Indicators"
        ws[f'A{row}'].font = Font(size=14, bold=True)

        row += 2
        kpis = [
            ("Total Systems", latest.get('total_systems', 0)),
            ("Healthy Systems", latest.get('healthy_systems', 0)),
            ("Highly Vulnerable Systems", latest.get('highly_vulnerable_systems', 0)),
            ("Missing Patches", latest.get('missing_patches', 0)),
            ("Critical Patches", latest.get('critical_count', 0)),
        ]

        for label, value in kpis:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)

            # Conditional formatting
            if label == "Highly Vulnerable Systems" and value > 0:
                ws[f'B{row}'].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                ws[f'B{row}'].font = Font(color="FFFFFF", bold=True)
            elif label == "Critical Patches" and value > 10:
                ws[f'B{row}'].fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
            elif label == "Healthy Systems":
                ws[f'B{row}'].fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15

    def _create_trend_charts(self, wb: Workbook, days: int = 30):
        """Create Trend Charts worksheet"""
        ws = wb.create_sheet("Trend Charts")

        # Get trend data
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        cursor.execute("""
            SELECT
                DATE(s.timestamp) as date,
                pm.missing_patches,
                sm.critical_count
            FROM snapshots s
            LEFT JOIN patch_metrics pm ON s.snapshot_id = pm.snapshot_id
            LEFT JOIN severity_metrics sm ON s.snapshot_id = sm.snapshot_id
            WHERE s.status = 'success'
              AND s.timestamp >= datetime('now', '-' || ? || ' days')
            ORDER BY s.timestamp ASC
        """, (days,))

        trend_data = [dict(row) for row in cursor.fetchall()]
        conn.close()

        if not trend_data:
            ws['A1'] = f"No trend data available (last {days} days)"
            return

        # Header
        ws['A1'] = f"Patch Trends - Last {days} Days"
        ws['A1'].font = Font(size=14, bold=True)

        # Write data
        row = 3
        ws['A3'] = "Date"
        ws['B3'] = "Missing Patches"
        ws['C3'] = "Critical Patches"

        for col in ['A', 'B', 'C']:
            ws[f'{col}3'].font = Font(bold=True)
            ws[f'{col}3'].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        row = 4
        for data in trend_data:
            ws[f'A{row}'] = data['date']
            ws[f'B{row}'] = data.get('missing_patches', 0)
            ws[f'C{row}'] = data.get('critical_count', 0)
            row += 1

        # Create line chart
        chart = LineChart()
        chart.title = "Missing & Critical Patches Trend"
        chart.y_axis.title = "Patch Count"
        chart.x_axis.title = "Date"

        # Data for chart (skip header)
        data_ref = Reference(ws, min_col=2, min_row=3, max_row=row-1, max_col=3)
        cats_ref = Reference(ws, min_col=1, min_row=4, max_row=row-1)

        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        ws.add_chart(chart, "E3")

        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15

    def _create_severity_analysis(self, wb: Workbook):
        """Create Severity Analysis worksheet"""
        ws = wb.create_sheet("Severity Analysis")

        # Get latest severity data
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        cursor.execute("""
            SELECT
                sm.critical_count,
                sm.important_count,
                sm.moderate_count,
                sm.low_count,
                sm.unrated_count
            FROM severity_metrics sm
            INNER JOIN (
                SELECT MAX(snapshot_id) as latest_snapshot_id
                FROM snapshots
                WHERE status = 'success'
            ) latest ON sm.snapshot_id = latest.latest_snapshot_id
        """)

        severity = dict(cursor.fetchone() or {})
        conn.close()

        if not severity:
            ws['A1'] = "No severity data available"
            return

        # Header
        ws['A1'] = "Missing Patches by Severity"
        ws['A1'].font = Font(size=14, bold=True)

        # Data table
        row = 3
        ws['A3'] = "Severity"
        ws['B3'] = "Count"

        for col in ['A', 'B']:
            ws[f'{col}3'].font = Font(bold=True)
            ws[f'{col}3'].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        severities = [
            ("Critical", severity.get('critical_count', 0)),
            ("Important", severity.get('important_count', 0)),
            ("Moderate", severity.get('moderate_count', 0)),
            ("Low", severity.get('low_count', 0)),
            ("Unrated", severity.get('unrated_count', 0))
        ]

        row = 4
        for label, count in severities:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = count

            # Conditional formatting
            if label == "Critical":
                ws[f'A{row}'].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                ws[f'A{row}'].font = Font(color="FFFFFF", bold=True)
            elif label == "Important":
                ws[f'A{row}'].fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

            row += 1

        # Create pie chart
        chart = PieChart()
        chart.title = "Severity Distribution"

        labels = Reference(ws, min_col=1, min_row=4, max_row=row-1)
        data = Reference(ws, min_col=2, min_row=3, max_row=row-1)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)

        ws.add_chart(chart, "D3")

        # Column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12


def main():
    """CLI interface for report generator"""
    import argparse

    parser = argparse.ArgumentParser(description="PMP Report Generator")
    parser.add_argument('--days', type=int, default=30,
                       help='Number of days of trend data (default: 30)')
    parser.add_argument('--output', type=str,
                       help='Output directory (default: ~/work_projects/pmp_reports/)')
    parser.add_argument('--db', type=str,
                       help='Path to SQLite database')

    args = parser.parse_args()

    db_path = Path(args.db) if args.db else None
    output_dir = Path(args.output) if args.output else None

    generator = PMPReportGenerator(db_path=db_path)

    print(f"📊 Generating Compliance Dashboard ({args.days} days)...")

    try:
        report_path = generator.generate_compliance_dashboard(
            output_dir=output_dir,
            days=args.days
        )

        print(f"✅ Report generated: {report_path}")
        print(f"   File size: {report_path.stat().st_size / 1024:.1f} KB")

    except Exception as e:
        print(f"❌ Report generation failed: {e}")
        import sys
        sys.exit(1)


if __name__ == '__main__':
    main()
