#!/usr/bin/env python3
"""
PMP Policy Analyzer - Phase 191
Comprehensive policy review, analysis, and recommendations

Features:
- Parse all deployment policies from JSON
- Analyze naming conventions, schedules, organizations
- Identify policy inconsistencies and risks
- Generate recommendations for standardization
- Export detailed Excel report

Author: Patch Manager Plus API Specialist Agent
Date: 2025-11-26
Version: 1.0 (Phase 191)
"""

import sqlite3
import json
import re
from pathlib import Path
from typing import Optional, List, Dict, Tuple
from datetime import datetime
from collections import defaultdict, Counter

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ Error: openpyxl not installed")
    print("Install with: pip3 install openpyxl")
    import sys
    sys.exit(1)


class PMPPolicyAnalyzer:
    """
    Comprehensive PMP policy analyzer and report generator

    Analyzes:
    - Deployment policies (naming, schedules, targeting)
    - Health policy configuration
    - Approval settings
    - Policy standardization opportunities
    """

    def __init__(self, db_path: Optional[Path] = None):
        """Initialize policy analyzer"""
        if db_path is None:
            db_path = Path.home() / ".maia/databases/intelligence/pmp_config.db"

        self.db_path = Path(db_path)

        if not self.db_path.exists():
            raise FileNotFoundError(f"Database not found: {self.db_path}")

        # Style definitions
        self.header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True, size=11)

        self.risk_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        self.warning_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        self.success_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        self.info_fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")

        self.border_thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def analyze_and_export(self, output_dir: Optional[Path] = None) -> Path:
        """
        Analyze all policies and generate comprehensive Excel report

        Returns:
            Path to generated Excel report
        """
        if output_dir is None:
            output_dir = Path.home() / "work_projects/pmp_reports"

        output_dir.mkdir(parents=True, exist_ok=True)

        print(f"\n{'='*70}")
        print(f"🔍 PMP POLICY ANALYSIS & RECOMMENDATIONS")
        print(f"{'='*70}\n")

        # Load policy data
        print("  → Loading policy data...")
        policies = self._load_policies()
        health_policy = self._load_health_policy()
        approval_settings = self._load_approval_settings()

        print(f"     Found: {len(policies)} deployment policies")

        # Analyze
        print("  → Analyzing policies...")
        analysis = self._analyze_policies(policies)

        print("  → Generating recommendations...")
        recommendations = self._generate_recommendations(analysis, health_policy, approval_settings)

        # Generate report
        timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        filename = f"PMP_Policy_Analysis_{timestamp}.xlsx"
        output_path = output_dir / filename

        print("  → Creating Excel report...")
        self._create_excel_report(output_path, policies, analysis, recommendations, health_policy, approval_settings)

        print(f"\n{'='*70}")
        print(f"✅ POLICY ANALYSIS COMPLETE")
        print(f"{'='*70}")
        print(f"📄 File: {output_path}")
        print(f"📊 Size: {output_path.stat().st_size / 1024:.1f} KB")
        print(f"{'='*70}\n")

        return output_path

    def _load_policies(self) -> List[Dict]:
        """Load all deployment policies from database"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        # Get latest snapshot that has policy data
        cursor.execute("""
            SELECT settings_json
            FROM deployment_policies
            WHERE snapshot_id = (
                SELECT MAX(snapshot_id)
                FROM deployment_policies
            )
        """)

        policies = []
        for row in cursor.fetchall():
            try:
                policy = json.loads(row[0])
                policies.append(policy)
            except json.JSONDecodeError:
                continue

        conn.close()
        return policies

    def _load_health_policy(self) -> Optional[Dict]:
        """Load health policy settings"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        cursor.execute("""
            SELECT settings_json
            FROM health_policy
            WHERE snapshot_id = (SELECT MAX(snapshot_id) FROM health_policy)
            LIMIT 1
        """)

        row = cursor.fetchone()
        conn.close()

        if row and row[0]:
            try:
                return json.loads(row[0])
            except json.JSONDecodeError:
                return None
        return None

    def _load_approval_settings(self) -> Optional[Dict]:
        """Load approval settings"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        cursor.execute("""
            SELECT settings_json
            FROM approval_settings
            WHERE snapshot_id = (SELECT MAX(snapshot_id) FROM approval_settings)
            LIMIT 1
        """)

        row = cursor.fetchone()
        conn.close()

        if row and row[0]:
            try:
                return json.loads(row[0])
            except json.JSONDecodeError:
                return None
        return None

    def _analyze_policies(self, policies: List[Dict]) -> Dict:
        """Analyze policies for patterns and issues"""
        analysis = {
            'total_policies': len(policies),
            'by_organization': defaultdict(list),
            'by_creator': Counter(),
            'naming_patterns': {
                'has_org_prefix': 0,
                'has_schedule': 0,
                'has_platform': 0,
                'naming_inconsistencies': []
            },
            'schedule_patterns': {
                'weekend': 0,
                'weekday': 0,
                'daily': 0,
                'unknown': 0
            },
            'unique_organizations': set(),
            'policy_age': {
                'recent': 0,  # < 3 months
                'moderate': 0,  # 3-12 months
                'old': 0  # > 12 months
            }
        }

        # Regex patterns
        org_pattern = re.compile(r'^([A-Z][A-Za-z\s&]+)\s*-')  # "GS1 - ..."
        schedule_keywords = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday',
                            'Weekend', 'Weekday', 'Daily', 'Weekly', 'Monthly']
        platform_keywords = ['Windows', 'Servers', 'Workstations', 'Mac', 'Linux']

        current_time = datetime.now().timestamp() * 1000  # ms

        for policy in policies:
            name = policy.get('template_name', 'Unnamed')
            creator = policy.get('first_name', 'Unknown')
            creation_time = policy.get('creation_time', 0)

            # Creator tracking
            analysis['by_creator'][creator] += 1

            # Extract organization from policy name
            org_match = org_pattern.match(name)
            if org_match:
                org = org_match.group(1).strip()
                analysis['unique_organizations'].add(org)
                analysis['by_organization'][org].append(name)
                analysis['naming_patterns']['has_org_prefix'] += 1
            else:
                analysis['naming_patterns']['naming_inconsistencies'].append(
                    f"No organization prefix: {name}"
                )

            # Schedule detection
            name_lower = name.lower()
            if any(kw.lower() in name_lower for kw in schedule_keywords):
                analysis['naming_patterns']['has_schedule'] += 1

                if 'sunday' in name_lower or 'saturday' in name_lower or 'weekend' in name_lower:
                    analysis['schedule_patterns']['weekend'] += 1
                elif any(day in name_lower for day in ['monday', 'tuesday', 'wednesday', 'thursday', 'friday']):
                    analysis['schedule_patterns']['weekday'] += 1
                elif 'daily' in name_lower:
                    analysis['schedule_patterns']['daily'] += 1
            else:
                analysis['schedule_patterns']['unknown'] += 1

            # Platform detection
            if any(kw.lower() in name_lower for kw in platform_keywords):
                analysis['naming_patterns']['has_platform'] += 1

            # Age calculation
            if creation_time > 0:
                age_ms = current_time - creation_time
                age_months = age_ms / (1000 * 60 * 60 * 24 * 30)

                if age_months < 3:
                    analysis['policy_age']['recent'] += 1
                elif age_months < 12:
                    analysis['policy_age']['moderate'] += 1
                else:
                    analysis['policy_age']['old'] += 1

        return analysis

    def _generate_recommendations(self, analysis: Dict, health_policy: Optional[Dict],
                                 approval_settings: Optional[Dict]) -> List[Dict]:
        """Generate recommendations based on analysis"""
        recommendations = []

        # Recommendation 1: Naming standardization
        if analysis['naming_patterns']['naming_inconsistencies']:
            recommendations.append({
                'priority': 'HIGH',
                'category': 'Naming Convention',
                'issue': f"{len(analysis['naming_patterns']['naming_inconsistencies'])} policies lack organization prefix",
                'recommendation': 'Standardize policy naming: [Organization] - [Platform] - [Schedule] - [Description]',
                'example': 'GS1 - Windows Servers - Weekend - Critical Patches',
                'affected_count': len(analysis['naming_patterns']['naming_inconsistencies'])
            })

        # Recommendation 2: Schedule clarity
        if analysis['schedule_patterns']['unknown'] > 0:
            recommendations.append({
                'priority': 'MEDIUM',
                'category': 'Schedule Clarity',
                'issue': f"{analysis['schedule_patterns']['unknown']} policies don't indicate schedule in name",
                'recommendation': 'Include deployment schedule in policy name for easier identification',
                'example': 'Add "Sunday 2AM" or "Weekday Evening" to policy names',
                'affected_count': analysis['schedule_patterns']['unknown']
            })

        # Recommendation 3: Platform specification
        total_with_platform = analysis['naming_patterns']['has_platform']
        total_policies = analysis['total_policies']
        if total_with_platform < total_policies * 0.8:
            recommendations.append({
                'priority': 'MEDIUM',
                'category': 'Platform Clarity',
                'issue': f"Only {total_with_platform}/{total_policies} policies specify platform",
                'recommendation': 'Clearly indicate target platform (Windows Servers, Workstations, Mac, Linux)',
                'example': 'GS1 - Windows Workstations - ...',
                'affected_count': total_policies - total_with_platform
            })

        # Recommendation 4: Health policy
        if health_policy:
            healthpolicy = health_policy.get('healthpolicy', {})
            if healthpolicy.get('consider_only_approved_patches'):
                recommendations.append({
                    'priority': 'INFO',
                    'category': 'Health Policy',
                    'issue': 'Health assessment only considers approved patches',
                    'recommendation': 'Ensure patch approval workflow is timely to avoid false "healthy" status',
                    'example': 'Systems may show "healthy" despite missing unapproved patches',
                    'affected_count': None
                })

        # Recommendation 5: Approval mode
        if approval_settings:
            approvalsettings = approval_settings.get('approvalsettings', {})
            if approvalsettings.get('patch_approval') == 'manual':
                recommendations.append({
                    'priority': 'HIGH',
                    'category': 'Approval Workflow',
                    'issue': 'Manual patch approval mode requires human intervention',
                    'recommendation': 'Consider automatic approval for low-risk patches (non-security, tested)',
                    'example': 'Auto-approve: Non-security updates, Office updates, tested patches',
                    'affected_count': None
                })

        # Recommendation 6: Old policies
        if analysis['policy_age']['old'] > 0:
            recommendations.append({
                'priority': 'LOW',
                'category': 'Policy Maintenance',
                'issue': f"{analysis['policy_age']['old']} policies are >12 months old",
                'recommendation': 'Review old policies for relevance and update deployment targets',
                'example': 'Verify old policies still target active systems',
                'affected_count': analysis['policy_age']['old']
            })

        # Recommendation 7: Organization coverage
        unique_orgs = len(analysis['unique_organizations'])
        recommendations.append({
            'priority': 'INFO',
            'category': 'Coverage Analysis',
            'issue': f"Policies cover {unique_orgs} organizations (vs 30 total in PMP)",
            'recommendation': 'Verify all organizations have appropriate deployment policies',
            'example': 'Check: Orro Group, Millennium Services, Wyvern Health, etc.',
            'affected_count': None
        })

        return recommendations

    def _create_excel_report(self, output_path: Path, policies: List[Dict], analysis: Dict,
                            recommendations: List[Dict], health_policy: Optional[Dict],
                            approval_settings: Optional[Dict]):
        """Create comprehensive Excel report"""
        wb = Workbook()
        wb.remove(wb.active)

        # Sheet 1: Executive Summary
        self._create_summary_sheet(wb, analysis, recommendations)

        # Sheet 2: All Policies
        self._create_policies_sheet(wb, policies)

        # Sheet 3: Recommendations
        self._create_recommendations_sheet(wb, recommendations)

        # Sheet 4: Policy Patterns
        self._create_patterns_sheet(wb, analysis)

        # Sheet 5: Health & Approval Settings
        self._create_settings_sheet(wb, health_policy, approval_settings)

        wb.save(output_path)

    def _create_summary_sheet(self, wb: Workbook, analysis: Dict, recommendations: List[Dict]):
        """Executive Summary sheet"""
        ws = wb.create_sheet("Executive Summary")

        # Header
        ws['A1'] = "PMP Policy Analysis - Executive Summary"
        ws['A1'].font = Font(size=16, bold=True, color="2F5496")
        ws.merge_cells('A1:D1')

        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(size=10, italic=True)
        ws.merge_cells('A2:D2')

        row = 4
        ws[f'A{row}'] = "Key Metrics"
        ws[f'A{row}'].font = Font(size=14, bold=True)

        row += 2
        metrics = [
            ("Total Deployment Policies", analysis['total_policies']),
            ("Organizations Covered", len(analysis['unique_organizations'])),
            ("Policy Creators", len(analysis['by_creator'])),
            ("Policies with Standard Naming", analysis['naming_patterns']['has_org_prefix']),
            ("Policies with Schedule Info", analysis['naming_patterns']['has_schedule']),
            ("Policies >12 Months Old", analysis['policy_age']['old']),
        ]

        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

        # Recommendations summary
        row += 2
        ws[f'A{row}'] = "Recommendations Summary"
        ws[f'A{row}'].font = Font(size=14, bold=True)

        row += 1
        high_priority = len([r for r in recommendations if r['priority'] == 'HIGH'])
        medium_priority = len([r for r in recommendations if r['priority'] == 'MEDIUM'])
        low_priority = len([r for r in recommendations if r['priority'] == 'LOW'])

        ws[f'A{row}'] = "High Priority"
        ws[f'B{row}'] = high_priority
        if high_priority > 0:
            ws[f'B{row}'].fill = self.risk_fill
            ws[f'B{row}'].font = Font(bold=True, color="FFFFFF")
        row += 1

        ws[f'A{row}'] = "Medium Priority"
        ws[f'B{row}'] = medium_priority
        if medium_priority > 0:
            ws[f'B{row}'].fill = self.warning_fill
            ws[f'B{row}'].font = Font(bold=True)
        row += 1

        ws[f'A{row}'] = "Low Priority"
        ws[f'B{row}'] = low_priority
        row += 1

        # Column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 15

    def _create_policies_sheet(self, wb: Workbook, policies: List[Dict]):
        """All Policies sheet"""
        ws = wb.create_sheet("All Policies")

        # Header
        ws['A1'] = f"All Deployment Policies ({len(policies)} policies)"
        ws['A1'].font = Font(size=14, bold=True)

        row = 3
        headers = ['Policy Name', 'Created By', 'Created Date', 'Modified Date', 'Set as Default']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border_thin

        row += 1
        for policy in policies:
            name = policy.get('template_name', 'Unnamed')
            creator = policy.get('first_name', 'Unknown')
            created = policy.get('creation_time', 0)
            modified = policy.get('modified_time', 0)
            is_default = policy.get('set_as_default', False)

            ws[f'A{row}'] = name
            ws[f'B{row}'] = creator
            ws[f'C{row}'] = datetime.fromtimestamp(created/1000).strftime('%Y-%m-%d') if created > 0 else 'N/A'
            ws[f'D{row}'] = datetime.fromtimestamp(modified/1000).strftime('%Y-%m-%d') if modified > 0 else 'N/A'
            ws[f'E{row}'] = 'Yes' if is_default else 'No'

            # Highlight default policies
            if is_default:
                ws[f'E{row}'].fill = self.success_fill
                ws[f'E{row}'].font = Font(bold=True)

            # Apply borders
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = self.border_thin

            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 70
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 14

    def _create_recommendations_sheet(self, wb: Workbook, recommendations: List[Dict]):
        """Recommendations sheet"""
        ws = wb.create_sheet("Recommendations")

        # Header
        ws['A1'] = f"Policy Recommendations ({len(recommendations)} items)"
        ws['A1'].font = Font(size=14, bold=True)

        row = 3
        headers = ['Priority', 'Category', 'Issue', 'Recommendation', 'Example', 'Affected']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border_thin

        row += 1
        for rec in recommendations:
            ws[f'A{row}'] = rec['priority']
            ws[f'B{row}'] = rec['category']
            ws[f'C{row}'] = rec['issue']
            ws[f'D{row}'] = rec['recommendation']
            ws[f'E{row}'] = rec['example']
            ws[f'F{row}'] = rec['affected_count'] if rec['affected_count'] is not None else 'N/A'

            # Priority color coding
            if rec['priority'] == 'HIGH':
                ws[f'A{row}'].fill = self.risk_fill
                ws[f'A{row}'].font = Font(bold=True, color="FFFFFF")
            elif rec['priority'] == 'MEDIUM':
                ws[f'A{row}'].fill = self.warning_fill
                ws[f'A{row}'].font = Font(bold=True)
            elif rec['priority'] == 'INFO':
                ws[f'A{row}'].fill = self.info_fill
                ws[f'A{row}'].font = Font(bold=True)

            # Apply borders
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = self.border_thin
                ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical='top')

            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 50
        ws.column_dimensions['F'].width = 10

        # Row heights for readability
        for r in range(4, row):
            ws.row_dimensions[r].height = 50

    def _create_patterns_sheet(self, wb: Workbook, analysis: Dict):
        """Policy Patterns sheet"""
        ws = wb.create_sheet("Policy Patterns")

        # Header
        ws['A1'] = "Policy Patterns & Analysis"
        ws['A1'].font = Font(size=14, bold=True)

        row = 3

        # Schedule patterns
        ws[f'A{row}'] = "Deployment Schedule Patterns"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1

        schedule_data = [
            ("Weekend Deployments", analysis['schedule_patterns']['weekend']),
            ("Weekday Deployments", analysis['schedule_patterns']['weekday']),
            ("Daily Deployments", analysis['schedule_patterns']['daily']),
            ("Unknown Schedule", analysis['schedule_patterns']['unknown']),
        ]

        for label, count in schedule_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = count
            row += 1

        row += 2

        # Policy age
        ws[f'A{row}'] = "Policy Age Distribution"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1

        age_data = [
            ("Recent (<3 months)", analysis['policy_age']['recent']),
            ("Moderate (3-12 months)", analysis['policy_age']['moderate']),
            ("Old (>12 months)", analysis['policy_age']['old']),
        ]

        for label, count in age_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = count
            row += 1

        row += 2

        # Top creators
        ws[f'A{row}'] = "Top Policy Creators"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1

        ws[f'A{row}'] = "Creator"
        ws[f'B{row}'] = "Policies Created"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        row += 1

        for creator, count in analysis['by_creator'].most_common(10):
            ws[f'A{row}'] = creator
            ws[f'B{row}'] = count
            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 15

    def _create_settings_sheet(self, wb: Workbook, health_policy: Optional[Dict],
                               approval_settings: Optional[Dict]):
        """Health & Approval Settings sheet"""
        ws = wb.create_sheet("Health & Approval")

        # Header
        ws['A1'] = "Health Policy & Approval Settings"
        ws['A1'].font = Font(size=14, bold=True)

        row = 3

        # Health Policy
        ws[f'A{row}'] = "Health Policy Configuration"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1

        if health_policy:
            healthpolicy = health_policy.get('healthpolicy', {})

            ws[f'A{row}'] = "Vulnerable Threshold"
            ws[f'B{row}'] = f"1+ important patches missing"
            row += 1

            ws[f'A{row}'] = "Highly Vulnerable Threshold"
            ws[f'B{row}'] = f"1+ critical patches missing"
            row += 1

            ws[f'A{row}'] = "Consider Only Approved Patches"
            ws[f'B{row}'] = "Yes" if healthpolicy.get('consider_only_approved_patches') else "No"
            if healthpolicy.get('consider_only_approved_patches'):
                ws[f'B{row}'].fill = self.warning_fill
                ws[f'B{row}'].font = Font(bold=True)
            row += 1
        else:
            ws[f'A{row}'] = "No health policy data available"
            row += 1

        row += 2

        # Approval Settings
        ws[f'A{row}'] = "Patch Approval Settings"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1

        if approval_settings:
            approvalsettings = approval_settings.get('approvalsettings', {})

            ws[f'A{row}'] = "Approval Mode"
            ws[f'B{row}'] = approvalsettings.get('patch_approval', 'Unknown').title()
            if approvalsettings.get('patch_approval') == 'manual':
                ws[f'B{row}'].fill = self.warning_fill
                ws[f'B{row}'].font = Font(bold=True)
            row += 1
        else:
            ws[f'A{row}'] = "No approval settings data available"
            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 40


def main():
    """CLI interface"""
    import argparse

    parser = argparse.ArgumentParser(description="PMP Policy Analyzer")
    parser.add_argument('--output', type=str,
                       help='Output directory (default: ~/work_projects/pmp_reports/)')
    parser.add_argument('--db', type=str,
                       help='Path to SQLite database')

    args = parser.parse_args()

    db_path = Path(args.db) if args.db else None
    output_dir = Path(args.output) if args.output else None

    try:
        analyzer = PMPPolicyAnalyzer(db_path=db_path)
        report_path = analyzer.analyze_and_export(output_dir=output_dir)

        print(f"✅ Open with: open '{report_path}'")

    except Exception as e:
        print(f"❌ Analysis failed: {e}")
        import traceback
        traceback.print_exc()
        import sys
        sys.exit(1)


if __name__ == '__main__':
    main()
