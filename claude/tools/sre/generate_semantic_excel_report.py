#!/usr/bin/env python3
"""
Generate Comprehensive Excel Report from Semantic Analysis

Creates detailed Excel workbook with multiple worksheets:
1. Executive Summary - High-level overview
2. All Tickets - Complete ticket dataset with embeddings metadata
3. Category Analysis - Breakdown by category
4. Root Cause Analysis - Breakdown by root cause
5. Account Analysis - Breakdown by client account
6. Category x Root Cause Matrix - Cross-tabulation
7. Timeline Analysis - Tickets over time
8. Sample Tickets - Representative examples per category

Created: 2025-10-27
Author: Maia SDM Agent
"""

import os
import sys
from pathlib import Path
from collections import Counter, defaultdict
from datetime import datetime
import sqlite3
import chromadb
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

MAIA_ROOT = Path(__file__).resolve().parents[3]
DB_PATH = MAIA_ROOT / "claude/data/servicedesk_tickets.db"

try:
    import openpyxl
except ImportError:
    print("❌ Missing dependency: openpyxl")
    print("   Install: pip3 install openpyxl")
    sys.exit(1)


def extract_keywords(text: str, top_n: int = 10) -> str:
    """Simple keyword extraction - returns comma-separated string"""
    stop_words = {
        'the', 'a', 'an', 'and', 'or', 'but', 'in', 'on', 'at', 'to', 'for',
        'of', 'with', 'by', 'from', 'as', 'is', 'was', 'are', 'were', 'been',
        'be', 'have', 'has', 'had', 'do', 'does', 'did', 'will', 'would',
        'should', 'could', 'may', 'might', 'must', 'can', 'this', 'that',
        'these', 'those', 'i', 'you', 'he', 'she', 'it', 'we', 'they',
        'what', 'which', 'who', 'when', 'where', 'why', 'how', 'issue',
        'resolution', 'ticket', 'user', 'client', 'customer'
    }

    words = text.lower().split()
    word_counts = Counter([
        word.strip('.,;:!?"\'()[]{}')
        for word in words
        if len(word) > 3 and word.strip('.,;:!?"\'()[]{}') not in stop_words
    ])

    return ', '.join([word for word, count in word_counts.most_common(top_n)])


def style_header(ws, row=1):
    """Apply header styling to worksheet"""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for cell in ws[row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def auto_adjust_columns(ws, max_width=50):
    """Auto-adjust column widths"""
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except Exception:
                pass

        adjusted_width = min(max_length + 2, max_width)
        ws.column_dimensions[column_letter].width = adjusted_width


# ========== Sheet Creation Helpers ==========

def _create_executive_summary_sheet(writer, df_all: pd.DataFrame) -> None:
    """Create Sheet 1: Executive Summary"""
    summary_data = {
        'Metric': [
            'Total Tickets Analyzed', 'Unique Categories', 'Unique Root Causes',
            'Unique Client Accounts', 'Date Range (Earliest)', 'Date Range (Latest)',
            'Tickets with Descriptions', 'Tickets with Solutions', 'Analysis Date',
            'Embedding Model Used'
        ],
        'Value': [
            f"{len(df_all):,}", f"{df_all['Category'].nunique():,}",
            f"{df_all['Root Cause'].nunique():,}", f"{df_all['Account Name'].nunique():,}",
            df_all['Created Time'].min() if df_all['Created Time'].min() else 'N/A',
            df_all['Created Time'].max() if df_all['Created Time'].max() else 'N/A',
            f"{len(df_all[df_all['Has Description'] == 'Yes']):,}",
            f"{len(df_all[df_all['Has Solution'] == 'Yes']):,}",
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'intfloat/e5-base-v2 (768-dimensional embeddings)'
        ]
    }
    df_summary = pd.DataFrame(summary_data)
    df_summary.to_excel(writer, sheet_name='Executive Summary', index=False)
    ws = writer.sheets['Executive Summary']
    style_header(ws)
    auto_adjust_columns(ws)


def _create_all_tickets_sheet(writer, df_all: pd.DataFrame) -> None:
    """Create Sheet 2: All Tickets (Complete Dataset)"""
    df_sorted = df_all.sort_values(['Category', 'Root Cause', 'Account Name'])
    df_sorted.to_excel(writer, sheet_name='All Tickets', index=False)
    ws = writer.sheets['All Tickets']
    style_header(ws)
    auto_adjust_columns(ws)
    ws.auto_filter.ref = ws.dimensions


def _create_category_analysis_sheet(writer, df_all: pd.DataFrame) -> None:
    """Create Sheet 3: Category Analysis"""
    category_analysis = []
    for category in df_all['Category'].value_counts().index:
        cat_tickets = df_all[df_all['Category'] == category]
        top_root_causes = cat_tickets['Root Cause'].value_counts().head(5)
        root_causes_str = ' | '.join([f"{rc} ({count})" for rc, count in top_root_causes.items()])
        top_accounts = cat_tickets['Account Name'].value_counts().head(3)
        accounts_str = ' | '.join([f"{acc} ({count})" for acc, count in top_accounts.items()])
        all_docs = ' '.join(cat_tickets['Document Excerpt'].head(50).tolist())
        keywords = extract_keywords(all_docs, top_n=15)

        category_analysis.append({
            'Category': category, 'Total Tickets': len(cat_tickets),
            'Percentage': f"{len(cat_tickets) / len(df_all) * 100:.1f}%",
            'Top Root Causes': root_causes_str, 'Top Affected Accounts': accounts_str,
            'Common Keywords': keywords,
            'Tickets with Solutions': len(cat_tickets[cat_tickets['Has Solution'] == 'Yes']),
            'Avg Resolution Rate': f"{len(cat_tickets[cat_tickets['Has Solution'] == 'Yes']) / len(cat_tickets) * 100:.1f}%"
        })

    df_category = pd.DataFrame(category_analysis).sort_values('Total Tickets', ascending=False)
    df_category.to_excel(writer, sheet_name='Category Analysis', index=False)
    ws = writer.sheets['Category Analysis']
    style_header(ws)
    auto_adjust_columns(ws)
    ws.auto_filter.ref = ws.dimensions


def _create_root_cause_analysis_sheet(writer, df_all: pd.DataFrame) -> None:
    """Create Sheet 4: Root Cause Analysis"""
    root_cause_analysis = []
    for root_cause in df_all['Root Cause'].value_counts().index:
        rc_tickets = df_all[df_all['Root Cause'] == root_cause]
        top_categories = rc_tickets['Category'].value_counts().head(5)
        categories_str = ' | '.join([f"{cat} ({count})" for cat, count in top_categories.items()])
        top_accounts = rc_tickets['Account Name'].value_counts().head(3)
        accounts_str = ' | '.join([f"{acc} ({count})" for acc, count in top_accounts.items()])
        all_docs = ' '.join(rc_tickets['Document Excerpt'].head(50).tolist())
        keywords = extract_keywords(all_docs, top_n=15)

        root_cause_analysis.append({
            'Root Cause': root_cause, 'Total Tickets': len(rc_tickets),
            'Percentage': f"{len(rc_tickets) / len(df_all) * 100:.1f}%",
            'Top Categories': categories_str, 'Top Affected Accounts': accounts_str,
            'Common Keywords': keywords
        })

    df_root_cause = pd.DataFrame(root_cause_analysis).sort_values('Total Tickets', ascending=False)
    df_root_cause.to_excel(writer, sheet_name='Root Cause Analysis', index=False)
    ws = writer.sheets['Root Cause Analysis']
    style_header(ws)
    auto_adjust_columns(ws)
    ws.auto_filter.ref = ws.dimensions


def _create_account_analysis_sheet(writer, df_all: pd.DataFrame) -> None:
    """Create Sheet 5: Account Analysis"""
    account_analysis = []
    for account in df_all['Account Name'].value_counts().head(50).index:
        acc_tickets = df_all[df_all['Account Name'] == account]
        top_categories = acc_tickets['Category'].value_counts().head(3)
        categories_str = ' | '.join([f"{cat} ({count})" for cat, count in top_categories.items()])
        top_root_causes = acc_tickets['Root Cause'].value_counts().head(3)
        root_causes_str = ' | '.join([f"{rc} ({count})" for rc, count in top_root_causes.items()])

        account_analysis.append({
            'Account Name': account, 'Total Tickets': len(acc_tickets),
            'Percentage': f"{len(acc_tickets) / len(df_all) * 100:.1f}%",
            'Top Categories': categories_str, 'Top Root Causes': root_causes_str,
            'Tickets with Solutions': len(acc_tickets[acc_tickets['Has Solution'] == 'Yes'])
        })

    df_account = pd.DataFrame(account_analysis).sort_values('Total Tickets', ascending=False)
    df_account.to_excel(writer, sheet_name='Account Analysis', index=False)
    ws = writer.sheets['Account Analysis']
    style_header(ws)
    auto_adjust_columns(ws)
    ws.auto_filter.ref = ws.dimensions


def _create_category_rootcause_matrix_sheet(writer, df_all: pd.DataFrame) -> None:
    """Create Sheet 6: Category x Root Cause Matrix"""
    pivot = pd.crosstab(df_all['Category'], df_all['Root Cause'], margins=True, margins_name='Total')
    pivot.to_excel(writer, sheet_name='Category x Root Cause')
    ws = writer.sheets['Category x Root Cause']

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.font = Font(bold=True)
    auto_adjust_columns(ws)


def _create_timeline_analysis_sheet(writer, df_all: pd.DataFrame) -> None:
    """Create Sheet 7: Timeline Analysis"""
    df_timeline = df_all.copy()
    df_timeline['Created Date'] = pd.to_datetime(df_timeline['Created Time'], errors='coerce')
    df_timeline['Year-Month'] = df_timeline['Created Date'].dt.to_period('M').astype(str)
    df_timeline = df_timeline[df_timeline['Year-Month'].notna()]

    if len(df_timeline) > 0:
        monthly_counts = df_timeline.groupby(['Year-Month', 'Category']).size().reset_index(name='Count')
        monthly_pivot = monthly_counts.pivot(index='Year-Month', columns='Category', values='Count').fillna(0)
        monthly_pivot['Total'] = monthly_pivot.sum(axis=1)
        monthly_pivot = monthly_pivot.sort_index()
        monthly_pivot.to_excel(writer, sheet_name='Timeline Analysis')
        ws = writer.sheets['Timeline Analysis']
        style_header(ws)
        auto_adjust_columns(ws)
    else:
        pd.DataFrame({'Note': ['No valid timestamps found in dataset']}).to_excel(
            writer, sheet_name='Timeline Analysis', index=False)


def _create_sample_tickets_sheet(writer, df_all: pd.DataFrame) -> None:
    """Create Sheet 8: Sample Tickets by Category"""
    sample_tickets = []
    for category in df_all['Category'].value_counts().head(20).index:
        cat_tickets = df_all[df_all['Category'] == category].head(5)
        for _, ticket in cat_tickets.iterrows():
            sample_tickets.append({
                'Category': category, 'Ticket ID': ticket['Ticket ID'],
                'Title': ticket['Title'], 'Account Name': ticket['Account Name'],
                'Root Cause': ticket['Root Cause'], 'Status': ticket['Status'],
                'Created Time': ticket['Created Time'],
                'Document Excerpt': ticket['Document Excerpt'][:300]
            })

    df_samples = pd.DataFrame(sample_tickets)
    df_samples.to_excel(writer, sheet_name='Sample Tickets', index=False)
    ws = writer.sheets['Sample Tickets']
    style_header(ws)
    auto_adjust_columns(ws)
    ws.auto_filter.ref = ws.dimensions


def _create_top_issues_sheet(writer, df_all: pd.DataFrame) -> None:
    """Create Sheet 9: Top Issues by Category"""
    top_issues = []
    for category in df_all['Category'].value_counts().head(10).index:
        cat_tickets = df_all[df_all['Category'] == category]
        title_counts = cat_tickets['Title'].value_counts().head(10)
        for title, count in title_counts.items():
            sample = cat_tickets[cat_tickets['Title'] == title].iloc[0]
            top_issues.append({
                'Category': category, 'Issue/Title': title[:100],
                'Occurrence Count': count, 'Root Cause': sample['Root Cause'],
                'Sample Ticket ID': sample['Ticket ID'], 'Has Solution': sample['Has Solution']
            })

    df_top_issues = pd.DataFrame(top_issues)
    df_top_issues = df_top_issues.sort_values(['Category', 'Occurrence Count'], ascending=[True, False])
    df_top_issues.to_excel(writer, sheet_name='Top Issues by Category', index=False)
    ws = writer.sheets['Top Issues by Category']
    style_header(ws)
    auto_adjust_columns(ws)
    ws.auto_filter.ref = ws.dimensions


def _create_automation_opportunities_sheet(writer, df_all: pd.DataFrame) -> None:
    """Create Sheet 10: Automation Opportunities"""
    automation_opps = []
    for category in df_all['Category'].value_counts().head(10).index:
        cat_tickets = df_all[df_all['Category'] == category]
        title_counts = cat_tickets['Title'].value_counts()
        repetitive_titles = title_counts[title_counts >= 5]

        if len(repetitive_titles) > 0:
            for title, count in repetitive_titles.head(20).items():
                sample_tickets = cat_tickets[cat_tickets['Title'] == title]
                solution_rate = len(sample_tickets[sample_tickets['Has Solution'] == 'Yes']) / len(sample_tickets) * 100
                most_common_rc = sample_tickets['Root Cause'].mode()[0] if len(sample_tickets['Root Cause'].mode()) > 0 else 'Unknown'

                automation_opps.append({
                    'Category': category, 'Repetitive Issue': title[:100],
                    'Occurrences': count, 'Root Cause': most_common_rc,
                    'Solution Rate': f"{solution_rate:.1f}%",
                    'Automation Potential': 'High' if count >= 20 and solution_rate >= 70 else 'Medium' if count >= 10 else 'Low',
                    'Sample Ticket IDs': ', '.join(sample_tickets['Ticket ID'].head(3).tolist())
                })

    df_automation = pd.DataFrame(automation_opps).sort_values('Occurrences', ascending=False)
    df_automation.to_excel(writer, sheet_name='Automation Opportunities', index=False)
    ws = writer.sheets['Automation Opportunities']
    style_header(ws)
    auto_adjust_columns(ws)
    ws.auto_filter.ref = ws.dimensions

    # Apply conditional formatting
    high_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    high_font = Font(color="006100")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=6):
        for cell in row:
            if cell.value == 'High':
                cell.fill = high_fill
                cell.font = high_font


def _load_chromadb_data():
    """Load data from ChromaDB collection"""
    rag_db_path = os.path.expanduser("~/.maia/servicedesk_semantic_analysis")
    client = chromadb.PersistentClient(path=rag_db_path)
    collection = client.get_collection(name="tickets_semantic")
    print(f"   Loading {collection.count():,} tickets from ChromaDB...")
    results = collection.get(include=['metadatas', 'documents'])
    return results['metadatas'], results['documents']


def _build_ticket_dataframe(metadatas: list, documents: list) -> pd.DataFrame:
    """Build master DataFrame from ChromaDB data"""
    all_tickets = []
    for idx, meta in enumerate(metadatas):
        all_tickets.append({
            'Ticket ID': meta.get('ticket_id', 'Unknown'),
            'Title': meta.get('title', ''),
            'Category': meta.get('category', 'Unknown'),
            'Status': meta.get('status', 'Unknown'),
            'Account Name': meta.get('account_name', 'Unknown'),
            'Created Time': meta.get('created_time', ''),
            'Closed Time': meta.get('closed_time', ''),
            'Assigned To': meta.get('assigned_to', ''),
            'Root Cause': meta.get('root_cause', 'Unknown'),
            'Has Description': 'Yes' if meta.get('has_description') else 'No',
            'Has Solution': 'Yes' if meta.get('has_solution') else 'No',
            'Document Excerpt': documents[idx][:500] if documents[idx] else ''
        })
    return pd.DataFrame(all_tickets)


def main():
    """Generate comprehensive Excel report from semantic analysis (refactored)"""
    print("📊 Generating Comprehensive Excel Report from Semantic Analysis...")

    # Load data
    metadatas, documents = _load_chromadb_data()
    print(f"   Loaded {len(metadatas):,} tickets")

    # Build DataFrame
    print("   Building master dataset...")
    df_all = _build_ticket_dataframe(metadatas, documents)

    # Create Excel workbook
    print("   Creating Excel workbook...")
    output_path = MAIA_ROOT / "claude/data/ServiceDesk_Semantic_Analysis_Report.xlsx"

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        print("   Creating Executive Summary sheet...")
        _create_executive_summary_sheet(writer, df_all)

        print("   Creating All Tickets sheet...")
        _create_all_tickets_sheet(writer, df_all)

        print("   Creating Category Analysis sheet...")
        _create_category_analysis_sheet(writer, df_all)

        print("   Creating Root Cause Analysis sheet...")
        _create_root_cause_analysis_sheet(writer, df_all)

        print("   Creating Account Analysis sheet...")
        _create_account_analysis_sheet(writer, df_all)

        print("   Creating Category x Root Cause Matrix...")
        _create_category_rootcause_matrix_sheet(writer, df_all)

        print("   Creating Timeline Analysis sheet...")
        _create_timeline_analysis_sheet(writer, df_all)

        print("   Creating Sample Tickets sheet...")
        _create_sample_tickets_sheet(writer, df_all)

        print("   Creating Top Issues by Category sheet...")
        _create_top_issues_sheet(writer, df_all)

        print("   Creating Automation Opportunities sheet...")
        _create_automation_opportunities_sheet(writer, df_all)

    print(f"\n✅ Excel report saved to: {output_path}")
    print(f"   File size: {output_path.stat().st_size / 1024:.1f} KB")
    print(f"\n📊 Worksheets created:")
    print(f"   1. Executive Summary - High-level metrics")
    print(f"   2. All Tickets - Complete dataset ({len(df_all):,} tickets)")
    print(f"   3. Category Analysis - {df_all['Category'].nunique()} categories")
    print(f"   4. Root Cause Analysis - {df_all['Root Cause'].nunique()} root causes")
    print(f"   5. Account Analysis - Top 50 accounts")
    print(f"   6. Category x Root Cause Matrix - Pivot table")
    print(f"   7. Timeline Analysis - Monthly trends")
    print(f"   8. Sample Tickets - Examples by category")
    print(f"   9. Top Issues by Category - Most common issues")
    print(f"   10. Automation Opportunities - High-volume repetitive tickets")


if __name__ == "__main__":
    main()
